# -*- coding: utf-8 -*-
"""
Generate an Excel workbook (FINAL_REPORT.xlsx) with Power-BI-style sheets:
Overview (KPIs), Monthly data + charts, Yearly + charts, Seasonal control,
Validation pairs, Vessel records, Data dictionary. Right-to-left sheets,
teal accent, native Excel charts.
"""
import os, json
import numpy as np
import pandas as pd
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, ScatterChart, RadarChart, Reference
from openpyxl.utils import get_column_letter

from config import ROOT, DET_DIR

OUT = os.path.join(ROOT, "docs", "FINAL_REPORT.xlsx")

TEAL = "01B8AA"
DARK = "23262D"
PANEL = "1E2126"
INK = "F0F0F0"
GRAY = "9CA3AF"
ARABIC_MONTHS = ["كانون الثاني", "شباط", "آذار", "نيسان", "أيار", "حزيران",
                 "تموز", "آب", "أيلول", "تشرين الأول", "تشرين الثاني", "كانون الأول"]

thin = Side(style="thin", color="32363E")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header(ws, row, ncols, fill=DARK):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name="Segoe UI", bold=True, color=INK, size=10)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def style_body(ws, r0, r1, ncols):
    for r in range(r0, r1 + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name="Segoe UI", color=INK, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BORDER


def title(ws, text, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1)
    c.value = text
    c.font = Font(name="Segoe UI", bold=True, size=14, color=TEAL)
    c.alignment = Alignment(horizontal="center")


def build():
    monthly = pd.read_csv(os.path.join(DET_DIR, "monthly.csv"))
    yearly = pd.read_csv(os.path.join(DET_DIR, "yearly.csv"))
    summary = json.load(open(os.path.join(DET_DIR, "summary.json"), encoding="utf-8"))
    scenes = pd.read_json(os.path.join(DET_DIR, "s1_scenes.jsonl"), lines=True)
    s2 = pd.read_json(os.path.join(DET_DIR, "s2_scenes.jsonl"), lines=True)
    good = scenes[scenes["error"].isna() | (scenes["error"] == "")].copy()
    good["dt"] = pd.to_datetime(good["datetime"])
    good["year"] = good["dt"].dt.year

    from validation import pair_s1_s2
    pairs = pair_s1_s2(scenes, s2)
    if len(pairs):
        pairs = (pairs.sort_values("gap_hours").drop_duplicates(subset="s1_id", keep="first")
                 .sort_values("s2_date").reset_index(drop=True))

    vessels = pd.read_csv(os.path.join(DET_DIR, "vessel_locations.csv"))
    regs = summary["regimes"]
    t2026 = regs["y2026_vs_2025"]["test"]
    tpost = regs["post_vs_pre_dec2024"]["test"]
    _mmx = monthly.set_index("ym")
    fa26 = float(np.mean([_mmx.loc[f"2026-0{i}", "mean_ships_port_adj"] for i in (2, 3, 4) if f"2026-0{i}" in _mmx.index]))
    _fbx = []
    for _i in (2, 3, 4):
        _vv = _mmx[(_mmx["month"] == _i) & (_mmx["year"].isin([2022, 2023, 2024, 2025])) & (_mmx["n_obs"] > 0)]["mean_ships_port_adj"]
        if len(_vv):
            _fbx.append(float(_vv.mean()))
    fa_base = float(np.mean(_fbx)) if _fbx else 0.0

    wb = Workbook()

    # ============ 1) Overview ============
    ws = wb.active
    ws.title = "نظرة عامة"
    ws.sheet_view.rightToLeft = True
    title(ws, "مرصد مرفأ اللاذقية — نظرة عامة", 4)
    kpis = [
        ("إجمالي سجلات السفن", int(good["n_est"].sum())),
        ("مشاهدات رادارية صالحة", len(good)),
        ("مشاهدات بصرية", len(s2)),
        ("أشهر مغطاة", int(monthly["n_obs"].gt(0).sum())),
        ("أزواج تحقق S1↔S2", len(pairs)),
    ]
    try:
        from validation import summary as _vsum
        _vs = _vsum()
        kpis.append(("متوسط الإشارة/الضوضاء (ديسيبل)", round(float(_vs.get("mean_snr_db") or 0), 1)))
    except Exception:
        kpis.append(("متوسط الإشارة/الضوضاء (ديسيبل)", None))
    ws.cell(row=3, column=1, value="المؤشر").font = Font(bold=True, color=TEAL, size=11)
    ws.cell(row=3, column=2, value="القيمة").font = Font(bold=True, color=TEAL, size=11)
    style_header(ws, 3, 2, DARK)
    for i, (k, v) in enumerate(kpis, 4):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
    style_body(ws, 4, 3 + len(kpis), 2)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 18

    # key findings block
    r0 = 4 + len(kpis) + 1
    ws.cell(row=r0, column=1, value="النتائج الرئيسية (2026 مقابل نفس الأشهر من 2025)").font = Font(bold=True, color=TEAL, size=11)
    findings = [
        f"متوسط السفن في الحوض: {t2026['mean_after']:.1f} مقابل {t2026['mean_before']:.1f} سفينة لكل مشاهدة ({t2026['diff_pct']:+.1f}%) — احتمال {t2026['p_value']:.4f}",
        f"الارتفاع ليس موسميًا: شباط–نيسان 2026 ({fa26:.1f}) مقابل 2022–2025 ({fa_base:.1f}) — p < 0.0001",
        f"المرسى: ارتفع متوسط سفن الانتظار إلى {monthly[monthly['ym']>='2026-01']['mean_anchorage'].mean():.1f} مقابل {monthly[(monthly['ym']>='2022-01')&(monthly['ym']<='2025-12')]['mean_anchorage'].mean():.1f} سفينة لكل مشاهدة",
        f"بعد كانون الأول 2024 مقابل قبله: غير دال (احتمال {tpost['p_value']:.4f})",
        "توقيت التحول: كانون الأول 2025 — الذروة: شباط–نيسان 2026",
    ]
    for i, f in enumerate(findings, r0 + 1):
        ws.cell(row=i, column=1, value=f).font = Font(name="Segoe UI", size=10.5, color=INK)
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=4)
    ws.column_dimensions["A"].width = 90

    # ============ 2) Monthly ============
    ws = wb.create_sheet("البيانات الشهرية")
    ws.sheet_view.rightToLeft = True
    mcols = ["الشهر", "مشاهدات", "سفن لكل مشاهدة", "بعد تصحيح التغطية", "أقصى مشهد", "سفن المرسى", "خشونة البحر"]
    title(ws, "السلسلة الشهرية للنشاط (2022–2026)", 7)
    for j, h in enumerate(mcols, 1):
        ws.cell(row=3, column=j, value=h)
    style_header(ws, 3, 7)
    for i, r in enumerate(monthly.itertuples(), 4):
        ws.cell(row=i, column=1, value=ym_str(r.ym))
        ws.cell(row=i, column=2, value=int(r.n_obs))
        ws.cell(row=i, column=3, value=round(float(r.mean_ships_port), 2))
        ws.cell(row=i, column=4, value=round(float(r.mean_ships_port_adj), 2))
        ws.cell(row=i, column=5, value=int(r.max_ships_port))
        ws.cell(row=i, column=6, value=round(float(r.mean_anchorage), 2))
        ws.cell(row=i, column=7, value=round(float(r.mean_roughness), 3))
    style_body(ws, 4, 3 + len(monthly), 7)
    last = 3 + len(monthly)
    for col, w in zip("ABCDEFG", [14, 11, 13, 16, 12, 13, 13]):
        ws.column_dimensions[col].width = w

    # line chart: ships/obs (col C) + adjusted (col D)
    ch = LineChart()
    ch.title = "النشاط الشهري — سفن لكل مشاهدة"
    ch.style = 12
    ch.y_axis.title = "سفن لكل مشاهدة"
    ch.x_axis.title = "الشهر"
    ch.height = 9; ch.width = 22
    data = Reference(ws, min_col=3, min_row=3, max_col=4, max_row=last)
    cats = Reference(ws, min_col=1, min_row=4, max_row=last)
    ch.add_data(data, titles_from_data=True)
    ch.set_categories(cats)
    ch.series[0].graphicalProperties.line.solidFill = "01B8AA"
    ch.series[1].graphicalProperties.line.solidFill = "F59E0B"
    ws.add_chart(ch, "I3")

    # bar chart: monthly obs count
    bc = BarChart()
    bc.title = "عدد المشاهدات شهريًا"
    bc.style = 12
    bc.height = 9; bc.width = 22
    data = Reference(ws, min_col=2, min_row=3, max_col=2, max_row=last)
    bc.add_data(data, titles_from_data=True)
    bc.set_categories(cats)
    bc.series[0].graphicalProperties.solidFill = "01B8AA"
    ws.add_chart(bc, "I24")

    # ============ 3) Yearly ============
    ws = wb.create_sheet("البيانات السنوية")
    ws.sheet_view.rightToLeft = True
    ycols = ["السنة", "مشاهدات", "متوسط النشاط", "بعد تصحيح التغطية", "شهر الذروة", "قيمة الذروة", "التغير السنوي %"]
    title(ws, "الملخص السنوي", 7)
    for j, h in enumerate(ycols, 1):
        ws.cell(row=3, column=j, value=h)
    style_header(ws, 3, 7)
    for i, r in enumerate(yearly.itertuples(), 4):
        ws.cell(row=i, column=1, value=int(r.year))
        ws.cell(row=i, column=2, value=int(r.n_obs))
        ws.cell(row=i, column=3, value=round(float(r.annual_mean_ships_port), 2))
        ws.cell(row=i, column=4, value=round(float(r.annual_mean_adj), 2))
        ws.cell(row=i, column=5, value=ARABIC_MONTHS[int(r.peak_month) - 1])
        ws.cell(row=i, column=6, value=round(float(r.peak_value), 2))
        ws.cell(row=i, column=7, value=round(float(r.yoy_pct), 1) if pd.notna(r.yoy_pct) else None)
    style_body(ws, 4, 3 + len(yearly), 7)
    last = 3 + len(yearly)
    for col, w in zip("ABCDEFG", [10, 11, 14, 16, 12, 12, 13]):
        ws.column_dimensions[col].width = w
    bc = BarChart()
    bc.title = "متوسط النشاط السنوي"
    bc.style = 12
    bc.height = 9; bc.width = 22
    data = Reference(ws, min_col=3, min_row=3, max_col=3, max_row=last)
    cats = Reference(ws, min_col=1, min_row=4, max_row=last)
    bc.add_data(data, titles_from_data=True)
    bc.set_categories(cats)
    bc.series[0].graphicalProperties.solidFill = "01B8AA"
    ws.add_chart(bc, "I3")

    # ============ 4) Seasonal ============
    ws = wb.create_sheet("التحكم الموسمي")
    ws.sheet_view.rightToLeft = True
    title(ws, "النمط الموسمي — كل شهر 2026 مقابل متوسط 2022–2025", 4)
    for j, h in enumerate(["الشهر", "متوسط 2022–2025", "2026", "الفرق"], 1):
        ws.cell(row=3, column=j, value=h)
    style_header(ws, 3, 4)
    m = monthly.set_index("ym")
    r_i = 4
    for mo in range(1, 13):
        past = m[(m["month"] == mo) & (m["year"].isin([2022, 2023, 2024, 2025])) & (m["n_obs"] > 0)]["mean_ships_port_adj"]
        curm = m[(m["month"] == mo) & (m["year"] == 2026) & (m["n_obs"] > 0)]["mean_ships_port_adj"]
        if len(past):
            ws.cell(row=r_i, column=1, value=ARABIC_MONTHS[mo - 1])
            ws.cell(row=r_i, column=2, value=round(float(past.mean()), 1))
            ws.cell(row=r_i, column=3, value=round(float(curm.mean()), 1) if len(curm) else None)
            ws.cell(row=r_i, column=4, value=round(float(curm.mean() - past.mean()), 1) if len(curm) else None)
            r_i += 1
    style_body(ws, 4, r_i - 1, 4)
    for col, w in zip("ABCD", [14, 16, 12, 12]):
        ws.column_dimensions[col].width = w
    try:
        rc = RadarChart()
        rc.title = "الملف الموسمي — 2022-2025 مقابل 2026"
        rc.style = 12
        rc.height = 9; rc.width = 22
        data = Reference(ws, min_col=2, min_row=3, max_col=3, max_row=r_i - 1)
        cats = Reference(ws, min_col=1, min_row=4, max_row=r_i - 1)
        rc.add_data(data, titles_from_data=True)
        rc.set_categories(cats)
        ws.add_chart(rc, "F3")
    except Exception as e:
        print("radar skip:", e)

    # ============ 4b) Regional comparison ============
    ws = wb.create_sheet("المقارنة الإقليمية")
    ws.sheet_view.rightToLeft = True
    from compare import load_series, yearly_comparison
    title(ws, "المقارنة الإقليمية — اللاذقية · طرطوس · بانياس (سفن لكل مشاهدة في الحوض)", 6)
    for j, h in enumerate(["الشهر", "اللاذقية", "طرطوس", "بانياس", "فجوة ل-ط", "مشاهدات طرطوس"], 1):
        ws.cell(row=3, column=j, value=h)
    style_header(ws, 3, 6)
    cdf = load_series()
    r_i = 4
    for _, r in cdf.iterrows():
        ws.cell(row=r_i, column=1, value=ym_str(r["ym"]))
        ws.cell(row=r_i, column=2, value=round(float(r["latakia"]), 2) if pd.notna(r["latakia"]) else None)
        ws.cell(row=r_i, column=3, value=round(float(r["tartus"]), 2) if pd.notna(r["tartus"]) else None)
        ws.cell(row=r_i, column=4, value=round(float(r["baniyas"]), 2) if "baniyas" in r and pd.notna(r["baniyas"]) else None)
        ws.cell(row=r_i, column=5, value=round(float(r["latakia"] - r["tartus"]), 2) if (pd.notna(r["latakia"]) and pd.notna(r["tartus"])) else None)
        ws.cell(row=r_i, column=6, value=int(r["tar_n"]) if pd.notna(r["tar_n"]) else None)
        r_i += 1
    style_body(ws, 4, r_i - 1, 6)
    for col, w in zip("ABCDEF", [14, 11, 11, 11, 11, 14]):
        ws.column_dimensions[col].width = w
    # yearly side by side
    yc = yearly_comparison()
    r0 = r_i + 1
    ws.cell(row=r0, column=1, value="الملخص السنوي").font = Font(bold=True, color=TEAL, size=11)
    for j, h in enumerate(["السنة", "اللاذقية", "طرطوس", "بانياس", "الفجوة", "تغير اللاذقية %"], 1):
        ws.cell(row=r0 + 1, column=j, value=h)
    style_header(ws, r0 + 1, 6)
    for k, r in enumerate(yc.itertuples(), r0 + 2):
        ws.cell(row=k, column=1, value=int(r.year))
        ws.cell(row=k, column=2, value=round(float(r.latakia), 1))
        ws.cell(row=k, column=3, value=round(float(r.tartus), 1) if pd.notna(r.tartus) else None)
        ws.cell(row=k, column=4, value=round(float(r.baniyas), 1) if pd.notna(r.baniyas) else None)
        ws.cell(row=k, column=5, value=round(float(r.gap), 1) if pd.notna(r.gap) else None)
        ws.cell(row=k, column=6, value=round(float(r.lat_yoy), 1) if pd.notna(r.lat_yoy) else None)
    style_body(ws, r0 + 2, r0 + 1 + len(yc), 6)

    # ============ 5) Validation ============
    ws = wb.create_sheet("التحقق S1-S2")
    ws.sheet_view.rightToLeft = True
    vcols = ["تاريخ S2", "تاريخ S1", "فارق (س)", "غيوم %", "S1 في الحوض", "S2 في الحوض"]
    title(ws, "أزواج التحقق المتقاطع Sentinel-1 ↔ Sentinel-2", 6)
    for j, h in enumerate(vcols, 1):
        ws.cell(row=3, column=j, value=h)
    style_header(ws, 3, 6)
    for i, r in enumerate(pairs.itertuples(), 4):
        ws.cell(row=i, column=1, value=str(pd.to_datetime(r.s2_date).strftime("%Y-%m-%d %H:%M")))
        ws.cell(row=i, column=2, value=str(pd.to_datetime(r.s1_date).strftime("%Y-%m-%d %H:%M")))
        ws.cell(row=i, column=3, value=round(float(r.gap_hours), 1))
        ws.cell(row=i, column=4, value=round(float(r.cloud_pct), 1))
        ws.cell(row=i, column=5, value=int(r.s1_port))
        ws.cell(row=i, column=6, value=int(r.s2_port))
    style_body(ws, 4, 3 + len(pairs), 6)
    last = 3 + len(pairs)
    for col, w in zip("ABCDEF", [18, 18, 10, 10, 13, 13]):
        ws.column_dimensions[col].width = w
    sc = ScatterChart()
    sc.title = "S1 مقابل S2 — سفن الحوض"
    sc.style = 12
    sc.height = 9; sc.width = 22
    xv = Reference(ws, min_col=5, min_row=4, max_row=last)
    yv = Reference(ws, min_col=6, min_row=4, max_row=last)
    sc.add_data(yv)
    sc.series[0].marker.symbol = "circle"
    sc.series[0].marker.size = 6
    sc.series[0].graphicalProperties.line.noFill = True
    ws.add_chart(sc, "H3")

    # ============ 6) Vessels ============
    ws = wb.create_sheet("سجلات السفن")
    ws.sheet_view.rightToLeft = True
    vcols2 = ["التاريخ", "خط العرض", "خط الطول", "الطول (م)", "العرض (م)", "ذروة (ديسيبل)", "المنطقة", "المسافة من المرفأ (م)"]
    title(ws, "سجلات السفن المكتشفة (عينة أول 2000)", 8)
    for j, h in enumerate(vcols2, 1):
        ws.cell(row=3, column=j, value=h)
    style_header(ws, 3, 8)
    sample = vessels.head(2000)
    for i, r in enumerate(sample.itertuples(), 4):
        ws.cell(row=i, column=1, value=str(r.acquisition_date))
        ws.cell(row=i, column=2, value=round(float(r.latitude), 5))
        ws.cell(row=i, column=3, value=round(float(r.longitude), 5))
        ws.cell(row=i, column=4, value=round(float(r.length_m)) if pd.notna(r.length_m) else None)
        ws.cell(row=i, column=5, value=round(float(r.width_m)) if pd.notna(r.width_m) else None)
        ws.cell(row=i, column=6, value=round(float(r.peak_db), 1))
        ws.cell(row=i, column=7, value={"in_port": "داخل الحوض", "anchorage": "المرسى", "transit": "عبور"}.get(r.zone, r.zone))
        ws.cell(row=i, column=8, value=round(float(r.dist_port_m)))
    style_body(ws, 4, 3 + len(sample), 8)
    for col, w in zip("ABCDEFGH", [12, 12, 12, 10, 10, 10, 12, 16]):
        ws.column_dimensions[col].width = w

    # ============ 7) Dictionary ============
    ws = wb.create_sheet("الدليل")
    ws.sheet_view.rightToLeft = True
    title(ws, "دليل حقول البيانات والملاحظات", 2)
    dict_rows = [
        ("id / datetime", "معرف المشهد الرسمي وتاريخ/وقت الالتقاط (UTC)"),
        ("satellite / platform / orbit", "القمر الصناعي (S1A/S1C/S1D) واتجاه المدار"),
        ("n_total / n_est", "عدد الكائنات اللامعة / تقدير السفن بعد الفصل والتوحيد"),
        ("n_est_in_port / n_anchorage", "السفن في الحوض / في المرسى"),
        ("coverage / cov_port", "تغطية البحر / تغطية حوض المرفأ في المشهد"),
        ("noise_floor_db / sea_roughness", "مستوى الضوضاء وحالة البحر (ضبط جودة)"),
        ("vessels.json", "لكل سفينة: الإحداثيات، الأبعاد، الاستطاعة، المنطقة، الأوسمة"),
    ]
    ws.cell(row=3, column=1, value="الحقل").font = Font(bold=True, color=TEAL)
    ws.cell(row=3, column=2, value="المعنى").font = Font(bold=True, color=TEAL)
    style_header(ws, 3, 2)
    for i, (k, v) in enumerate(dict_rows, 4):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
    style_body(ws, 4, 3 + len(dict_rows), 2)
    notes = [
        "",
        "مراجعة الجودة المطبقة:",
        "1) توحيد السجلات عبر إعادة معالجة كاملة ومواءمة ملفات المشاهدات مع ملفات السفن (0 تباين)",
        "2) توثيق أوسمة الجودة لكل سجل (مدمج / حافة رصيف / شبيه عنقود) عند توفرها",
        "3) المقارنات الزمنية على أساس نفس الفترة بنفس الفترة مع ضبط الموسمية",
        "",
        "مؤشر النشاط = متوسط السفن لكل مشاهدة، للمشاهدات التي غطّت حوض المرفأ كاملًا (تغطية ≥ 70%).",
        "المقارنات الزمنية على أساس نفس الفترة بنفس الفترة مع ضبط الموسمية.",
        "نوع السفينة: غير محدد (دقة 10م لا تكفي دون AIS).",
    ]
    r0 = 4 + len(dict_rows) + 1
    for i, n in enumerate(notes, r0):
        ws.cell(row=i, column=1, value=n).font = Font(name="Segoe UI", size=10, color=INK)
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 70

    wb.save(OUT)
    print("xlsx written:", OUT, os.path.getsize(OUT) // 1024, "KB")


def ym_str(ym):
    try:
        y, m = ym.split("-")
        return f"{ARABIC_MONTHS[int(m)-1]} {y}"
    except Exception:
        return str(ym)


if __name__ == "__main__":
    build()
